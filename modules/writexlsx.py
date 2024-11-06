# -*- coding: utf-8 -*-

"""
Archerank2

Copyright (C) <2018-2024> Markus Hackspacher

This file is part of Archerank2.

Archerank2 is free software: you can redistribute it and/or modify
it under the terms of the GNU General Public License as published by
the Free Software Foundation, either version 3 of the License, or
(at your option) any later version.

Archerank2 is distributed in the hope that it will be useful,
but WITHOUT ANY WARRANTY; without even the implied warranty of
MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
GNU Lesser General Public License for more details.

You should have received a copy of the GNU General Public License
along with Archerank2.  If not, see <http://www.gnu.org/licenses/>.
"""

import_openpyxl = True
try:
    from openpyxl import Workbook
except ImportError:
    print('Not found openpyxl, no export xlsx possible')
    import_openpyxl = False


class writexlsx:
    def __init__(self):
        self.wb = Workbook()
        self.wsWinner = self.wb.active
        self.wsWinner.title = "winner"
        self.wsUser = self.wb.create_sheet("user")
        self.wsAddress = self.wb.create_sheet("club")
        self.wsBow = self.wb.create_sheet("bow")
        self.wsAge = self.wb.create_sheet("age")

    def save(self, filename):
        self.wb.save(filename)

    def winner(self, data):
        """write winner data in sheet"""
        self.wsWinner.append(data)

    def user(self, data):
        """write user data in sheet"""
        self.wsUser.append(data)

    def adresse(self, data):
        """write adresse data in sheet"""
        self.wsAddress.append(data)

    def bow(self, data):
        """write Bow data in sheet"""
        self.wsBow.append(data)

    def age(self, data):
        """write age data in sheet"""
        self.wsAge.append(data)
